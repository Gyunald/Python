from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import streamlit as st
import pandas as pd
from datetime import time, datetime, timedelta
import qrcode
from streamlit_server_state import server_state, server_state_lock


st.set_page_config(page_title="IMI e-MES",layout='centered',initial_sidebar_state='collapsed') # layout='wide','centered', initial_sidebar_state='collapsed','expanded'

if 'direct' not in st.session_state:
    st.session_state.direct = pd.DataFrame(columns=['샵오더','OP No','상세내용','작업코드','시작시간', '종료시간','수량', '비고'])

if 'messages' not in st.session_state:
    st.session_state.messages = []

if 'overtime_data' not in st.session_state:
    st.session_state.overtime_data = pd.DataFrame()

def run():
    chrome_optiins = Options() # 브라우저 꺼짐 방지
    chrome_optiins.add_experimental_option('detach', True)
    chrome_optiins.add_experimental_option('excludeSwitches', ['enable-logging']) # 불필요한 에러 메세지 삭제
    service = Service(executable_path=ChromeDriverManager().install()) # 크롬 드라이버 최신 버전 자동 설치 후 서비스 만들기
        
    #URL 얻기
    driver = webdriver.Chrome(options= chrome_optiins, service=service)

    # 크롬 브라우저 사이즈
    driver.minimize_window()
    # driver.set_window_size(1400, 1000)

    # 주소 이동
    driver.get('https://emes.imi-critical.com/imi_clive/')

    driver.find_element(By.NAME, 'log_id').send_keys(id)
    driver.find_element(By.NAME, 'passwd').send_keys(pw)

    # # ID PW 저장
    # driver.find_element(By.XPATH, '//*[@id="save_id"]').click()
    # driver.find_element(By.XPATH, '//*[@id="save_pw"]').click()
    
    driver.find_element(By.XPATH, '//*[@id="btLogin"]').click()
    
    # 5초 까지 대기
    driver.implicitly_wait(5)

    # # 현재창 인식
    # driver.switch_to.frame(driver.find_element(By.XPATH,'//*[@id="frMain"]'))

    # # 작업일보 이동
    # driver.find_element(By.XPATH, '//*[@id="div_main_menu"]/ul[1]/li[4]/a').click()
    # driver.find_element(By.XPATH, '//*[@id="div_main_menu"]/ul[1]/li[4]/ul/li[5]/a').click()

    driver.get('https://emes.imi-critical.com/imi_clive/page/process_plan/work_result_disp.php')


    # 5초 까지 대기
    driver.implicitly_wait(5)

    # 작업일보 작성
    driver.find_element(By.XPATH, '//*[@id="div_top"]/div[1]/button[2]').click()

    # 5초 까지 대기
    driver.implicitly_wait(5)

    # 팀선택
    driver.find_element(By.XPATH, '//*[@id="div_pop_result"]/div[2]/table/tbody/tr/td[2]/select').click()
    if name == '박민호':
        driver.find_element(By.XPATH, '//*[@id="div_pop_result"]/div[2]/table/tbody/tr/td[2]/select/option[6]').click()
    else:
        driver.find_element(By.XPATH, '//*[@id="div_pop_result"]/div[2]/table/tbody/tr/td[2]/select/option[4]').click()

    # 작업자선택
    driver.find_element(By.XPATH, f"//select[@name='s_WR_Worker']/option[text()='{name}']").click()

    for i in range(1,len(st.session_state.direct)+1):
        if st.session_state.direct['작업코드'][i-1] == 'ID':
            driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[5]/select/option[9]').click()
            driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[6]/input').send_keys(st.session_state.direct['상세내용'][i-1])
            driver.find_element(By.XPATH, f"//*[@id='id_{i}']/td[7]/select/option[text()='{st.session_state.direct['작업코드'][i-1]}']").click()
            driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[9]/input[1]').send_keys(st.session_state.direct['시작시간'][i-1])
            driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[10]/input[1]').send_keys(st.session_state.direct['종료시간'][i-1])
            continue
       
        if name != '김진석' :
            # 샵오더
            driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[3]/input').send_keys(st.session_state.direct['샵오더'][i-1])
            driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[3]/input').send_keys(Keys.TAB)
            # OP No
            driver.find_element(By.XPATH, f"//*[@id='id_{i}']/td[4]/select/option[text()='{st.session_state.direct['OP No'][i-1]}']").click()

        # 작업명
        driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[5]/select/option[11]').click()

        # 상세내용        
        driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[6]/input').send_keys(st.session_state.direct['상세내용'][i-1])
        
        # 작업코드
        if st.session_state.direct['작업코드'][i-1] == 'F':
            driver.find_element(By.XPATH, f"//*[@id='id_{i}']/td[7]/select/option[text()='{st.session_state.direct['작업코드'][i-1]}']").click()
        # 수량
            driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[12]/input').send_keys(st.session_state.direct['수량'][i-1])
        else:
            driver.find_element(By.XPATH, f"//*[@id='id_{i}']/td[7]/select/option[text()='{st.session_state.direct['작업코드'][i-1]}']").click()

        # 비고
        driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[13]/input').send_keys(st.session_state.direct['비고'][i-1])

        driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[9]/input[1]').send_keys(st.session_state.direct['시작시간'][i-1])
        driver.find_element(By.XPATH, f'//*[@id="id_{i}"]/td[10]/input[1]').send_keys(st.session_state.direct['종료시간'][i-1])

    # # 등록
    driver.find_element(By.XPATH, '//*[@id="div_pop_result"]/div[4]/button').click()
    driver.find_element(By.XPATH, '/html/body/div[4]/div/table/tr[2]/td/button[1]').click()
    driver.close()

def direct_schedule(start_time, end_time, s2='', s3=''):
    schedule = st.session_state.direct
    breaks = [s2, start_time, '10:00', '10:10', '11:45', '12:45', '15:00', '15:20', '17:00', '17:20', s3, end_time]
    breaks = sorted(list(set(breaks)))

    if s2 < end_time:
        for i in range(len(breaks) - 1):
            if breaks[i] == '11:45' or breaks[i] == '17:00':
                continue
            elif start_time < breaks[i + 1] <= end_time:
                schedule = pd.concat([pd.DataFrame([{'샵오더' : so, 'OP No' : op,  '상세내용' : work,'작업코드': 'I', '시작시간': breaks[i], '종료시간': breaks[i + 1], '수량':'', '비고' : etc}]), schedule])
                if breaks[i+1] == end_time and finish == True:
                    schedule = pd.concat([pd.DataFrame([{'샵오더' : so, 'OP No' : op,  '상세내용' : work,'작업코드': 'F', '시작시간': breaks[i], '종료시간': breaks[i + 1], '수량': 수량, '비고' : etc}]), schedule])
                    
                if breaks[i] == '10:00' or breaks[i] == "15:00" :
                    schedule = pd.concat([pd.DataFrame([
                        {'상세내용' : '휴식', '작업코드': 'ID','시작시간': breaks[i], '종료시간': breaks[i + 1], }
                        ]), schedule])

                elif s2 == breaks[i] and s3 == breaks[i+1]:
                    schedule = pd.concat([pd.DataFrame([{'상세내용' : indirect, '작업코드': 'ID', '시작시간': breaks[i], '종료시간': breaks[i+1], }]), schedule])

        schedule=schedule.drop_duplicates(['시작시간','종료시간']).sort_values('시작시간').reset_index(drop=True)
        st.session_state.direct = schedule
        return schedule
    else:
        st.error('시간 확인')

def overtime():
    with server_state_lock.ot:
        schedule1 = server_state.ot
        schedule1.loc[schedule1[schedule1['이름'] == name].index, '상세내용'] = st.session_state.ot
        schedule1.loc[schedule1[schedule1['이름'] == name].index, '계획'] = st.session_state.r1
        server_state.ot = schedule1
        st.experimental_rerun()

def update_overtime_excel():
    file_path = 'C:/Users/kdkim/Desktop/overtime2.xlsx'
    current_month = datetime.now().month
    workbook = load_workbook(file_path)
    worksheet = workbook[f"{current_month}월"]
    target_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    target_part = 'Machining'

    found_row = None
    found_col = None

    for row in worksheet.iter_rows():
        if found_row is not None:
            break

        for cell in row:
            if cell.value == target_part:
                found_row = cell
                break
            elif cell.value == target_date:
                found_col = cell
                break

    row_idx = found_row.row
    col_idx = found_col.column

    font = Font(name='Arial', bold=False, italic=False, color='FF000000', size=10)

    for index, row_data in st.session_state.overtime_data.iterrows():
        for col_index, value in enumerate(row_data):
            cell = worksheet.cell(row=row_idx + index, column=col_idx + col_index, value=value)
            cell.font = font

    workbook.save('C:/Users/kdkim/Desktop/overtime3.xlsx')
    workbook.close()
current_date = datetime.now().date()
st.title(current_date)

employee_names = ['김진석', '이병호', '최재형', '김규덕', '이호성', '김태훈', '이영석', '임대건', '안형철', '허남윤', '김정훈', '황범식', '육신현', '정대영', '박민호']

current_week = datetime.today().isocalendar()[1]
if current_week % 2 == 0:
    team_members = ['이영석', '김정훈', '황범식']
else:
    team_members = ['임대건', '허남윤', '정대영']


t1,t2,t3= st.tabs(['작업일보','잔업', '히트넘버 & QR코드'])

with t1 :
    c1, c2, c3, = st.columns(3)

    with c1:
        name = st.selectbox('이름',sorted(employee_names),key='name')
        so = st.text_input('샵오더')
        op = st.selectbox('OP', ['OP 10','OP 20', 'OP 30', 'OP 40', 'OP 50', 'OP 60'], index=1)
        if name == '김진석' :
            work = st.text_input('상세내용','공정진행')
            
        elif name == '안형철':
            work = st.selectbox('상세내용',['MC3-6500','MC3-9500'])
        elif name == '박민호':
            work = st.selectbox('상세내용',['3D프린트 준비작업&블래스팅작업', '3D 프린트 언팩킹', '방전 가공','기타'])
        else:
            work = st.selectbox('상세내용',['세팅','황삭','정삭','수정','기타'], index=1)
        if work == '기타':
            work = st.text_input('기타',key='work')

        finish = st.checkbox('가공완료')
        if finish :
            수량 = str(st.number_input('수량',1,step=1))

        indirect_check = st.checkbox('간접작업')

    with c2:
        start_time = st.time_input('시작 시간', time(8,0), step=600)
        end_time = st.time_input('종료 시간', time(17,0), step=600)

        
        etc = st.text_input('비고')
        id ='kdkim'
        pw = 'xjsld12#'    
        add_button = st.button('추가', use_container_width=True)
        botton_run = st.button('등록', use_container_width=True , type= 'primary')

    with c3:
        if indirect_check:
            start_time2 = st.time_input('간접작업 시작', time(8, 0), step=600)
            end_time2 = st.time_input('간접작업 종료',
                                    (datetime.combine(datetime.min, start_time2) + timedelta(minutes=30)).time(),
                                    step=600)
            if name == '김진석' :
                indirect = st.selectbox('간접작업', ['생산회의', 'BOM 도면배포','타운홀미팅', '직무교육', '기타'])
            else:
                indirect = st.selectbox('간접작업', ['타운홀미팅', '직무교육', '회의','청소', '기타'])
            
            if indirect == '기타':
                indirect = st.text_input('기타', key='indirect')            
        
            if add_button:
                new_schedule = direct_schedule(start_time.strftime('%H:%M'), end_time.strftime('%H:%M'), start_time2.strftime('%H:%M'), end_time2.strftime('%H:%M'))
                st.session_state.direct = pd.concat([st.session_state.direct, new_schedule]).drop_duplicates().sort_values(['시작시간','종료시간']).reset_index(drop=True)
                
                for j in st.session_state.direct.loc[st.session_state.direct['시작시간'].duplicated() == True].index:
                    if j in st.session_state.direct.index:
                        st.session_state.direct.drop(index=j, inplace=True)
                
                for k in st.session_state.direct[st.session_state.direct['시작시간'] < st.session_state.direct['종료시간'].shift(1)].index:
                    if k in st.session_state.direct.index:
                        st.session_state.direct.loc[k-1, '종료시간'] = st.session_state.direct.loc[k, '시작시간']

        else:
            if add_button:
                new_schedule = direct_schedule(start_time.strftime('%H:%M'), end_time.strftime('%H:%M'))
                st.session_state.direct = pd.concat([st.session_state.direct, new_schedule]).drop_duplicates().sort_values(['시작시간','종료시간']).reset_index(drop=True)
                
                for j in st.session_state.direct.loc[st.session_state.direct['시작시간'].duplicated() == True].index:
                    if j in st.session_state.direct.index:
                        st.session_state.direct.drop(index=j, inplace=True)
                
                for k in st.session_state.direct[st.session_state.direct['시작시간'] < st.session_state.direct['종료시간'].shift(1)].index:
                    if k in st.session_state.direct.index:
                        st.session_state.direct.loc[k-1, '종료시간'] = st.session_state.direct.loc[k, '시작시간']

                st.toast('등록은 한 번만 누르세요', icon ='🙏')

    st.session_state.direct = st.session_state.direct.reset_index(drop=True)
    st.dataframe(st.session_state.direct.reset_index(drop=True).fillna(''),use_container_width=True,hide_index=True)
    
    if botton_run:
        st.toast('잘못 눌렀으면 새로고침', icon ='🙏')
        with c3:
            try:
                with st.spinner('🎈 업데이트 중'):
                    run()
                st.write('###')
                st.info('등록 완료 🙉')
                st.toast('수정 or 삭제는 홈페이지에서', icon ='💡')

            except:
                with c3:
                    st.write('###')
                    a = [st.session_state.direct['샵오더'][i] for i in st.session_state.direct[st.session_state.direct['작업코드'] == 'F'].index]
                    st.error(f"""
                    등록 실패 🙈\n
                    잔여수량 가공완료 확인\n
                    {' , '.join(a)}
                    """)

with t2:
    file_path = 'C:/Users/kdkim/Desktop/overtime.csv'
    overtime_data_csv = pd.read_csv(file_path, header=6)
    overtime_data_csv = overtime_data_csv[overtime_data_csv['공정명'] == 'Machining'].dropna(axis=1, how='all').fillna('')
    overtime_data_csv.rename(columns=lambda x: str(x).split('00:00:00')[0].replace('2023-', '').strip(), inplace=True)
    today_date = datetime.now().strftime('%m-%d')

    default_dataframe = pd.DataFrame({'이름': employee_names, '상세내용': ''})

    for i, value in enumerate(overtime_data_csv.iloc[:, overtime_data_csv.columns.get_loc(today_date) + 1]):
        if value != '':
            default_dataframe['상세내용'].iloc[i] = value
        else:
            default_dataframe['상세내용'].iloc[i] = overtime_data_csv.iloc[:, overtime_data_csv.columns.get_loc(today_date) - 1].iloc[i]

    with server_state_lock.ot:
        if "overtime_data" not in server_state:
            server_state.overtime_data = default_dataframe

    t2_c1, t2_c2 = st.columns([.7,.3])
    with t2_c1:
        edited_dataframe = st.data_editor(
            server_state.overtime_data,
            use_container_width=True,
            hide_index=True,
            height=563,
            disabled=['이름']
        )
    edited_dataframe['상세내용'] = edited_dataframe['상세내용'].fillna('')
    st.session_state.overtime_data = edited_dataframe

    leave_types = '휴가', '건강검진', '휴직'
    leave_types_pattern = '|'.join(leave_types)
    shift_am = '오전'
    shift_pm = '오후'

    for index, row in st.session_state.overtime_data.iterrows():
        if row['이름'] in team_members:
            st.session_state.overtime_data.at[index, '계획'] = '2교대'

        elif row['상세내용'] == '':
            st.session_state.overtime_data.at[index, '계획'] = '1'
            
        else:
            st.session_state.overtime_data.at[index, '계획'] = '2'
            st.session_state.overtime_data.loc[st.session_state.overtime_data['상세내용'].str.contains(leave_types_pattern), '계획'] = 'A'
            st.session_state.overtime_data.loc[st.session_state.overtime_data['상세내용'].str.contains(shift_am), '계획'] = 'B'
            st.session_state.overtime_data.loc[st.session_state.overtime_data['상세내용'].str.contains(shift_pm), '계획'] = 'C'

    st.session_state.overtime_data = st.session_state.overtime_data.reindex(columns=['이름', '계획', '상세내용'])
    with t2_c2:
        update_button_clicked = st.button('업데이트', use_container_width=True)
    
        if update_button_clicked:            
            with server_state_lock.ot:
                server_state.overtime_data = st.session_state.overtime_data
                
            st.session_state.overtime_data = st.session_state.overtime_data.reindex(columns=['계획', '상세내용'])
            with st.spinner('등록중...',):
                update_overtime_excel()
            st.info('업데이트 완료')

    # if st.button('초기화'):
    #     server_state.clear()

with t3:
    def generate_qrcode(number):
            qr = qrcode.make(number).resize((150, 150))
            return qr

    def convert_to_special_char(char):
        key_mapping = {
            '0': ')',
            '1': '!',
            '2': '@',
            '3': '#',
            '4': '$',
            '5': '%',
            '6': '^',
            '7': '&',
            '8': '*',
            '9': '(',
            '`': '~',
            '-': '_',
            '=': '+'
        }
        return key_mapping.get(char, char)

    def convert_string_to_special(input_string):
        return ''.join(convert_to_special_char(char) for char in input_string)

    path1 = "C:/Users/kdkim/Desktop/PO23.csv"
    path2 = "C:/Users/kdkim/Desktop/OSM23.csv"
    a1 = pd.read_csv(path1, header=9,)
    a2 = pd.read_csv(path2, header=10,)

    
    df = pd.merge(a1[['Part Name','Material','Heat No.']], a2[['Serial no.','Part Name','Part Number']], on='Part Name', how='outer').drop_duplicates().fillna('')
    df.columns = df.columns.str.replace('.','').str.strip()
    
    pn = st.selectbox('시리얼넘버',sorted([i for i in df['Serial no'].drop_duplicates()]),placeholder='검색 또는 선택')
    df = df[df['Serial no'].str.contains(pn,case=False)].reset_index(drop=True)
    df['QRcode'] = False

    if pn :

        word = ['ASTM/ASME','ASTM','ASME', 'A/S','SA/','/','Type1',]
        for i in word:
            df['Material'] = df['Material'].str.replace(i,'').str.strip()
        for j in ',;:':
            df['Part Name'] = df['Part Name'].str.strip().str.split(j).str[0].str.upper()
        df = df[['Part Name','Serial no','Material','Heat No', 'Part Number',"QRcode"]]
        df.rename(columns={'Serial no' : 'Serial No'}, inplace=True)

        t3_c = st.empty()                
        
        q = st.data_editor(
            df.reindex(columns=['QRcode','Part Name', 'Serial No', 'Material', 'Heat No', 'Part Number']),
            key='q',
            use_container_width=True,
            hide_index=True

            )

        index = q.loc[q['QRcode'] == True].index
        
        if len(index) == 1:
            with t3_c:
                t3_c2 = st.columns(7)
                for i in q.loc[index].values:
                    for j,k in zip(i[2:],range(0,len(c2),2)):
                        with c2[k]:
                            k = q.loc[index].columns[k//2 + 1]
                            if j :
                                barcode_image = generate_qrcode(convert_string_to_special(j.upper()))
                                st.image(barcode_image, caption=k, use_column_width=True)
        elif len(index) > 1:
            with t3_c:                        
                st.info('QRcode는 한 개씩만 생성됩니다')